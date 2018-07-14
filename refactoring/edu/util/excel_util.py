from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from decimal import *
from openpyxl import Workbook
import os


# 创建一个表格
def createExcel():
    wb = Workbook()
    return wb;


# 创建一个sheet
def creatSheet(wb, isFirst, sheetName):
    print('创建一个sheet--->%s' % (sheetName))
    if isFirst:
        ws = wb.active
    else:
        ws = wb.create_sheet()
    ws.title = sheetName
    return ws


# 把rows写进sheet
def putRowsToSheet(rows, ws):
    for row in rows:
        ws.append(row)
    adjustAlignmentAndColumnWidth(rows, ws)


# 居中和自动调整列宽
def adjustAlignmentAndColumnWidth(rows, ws):
    column_widths = []
    for j, row in enumerate(rows):
        for i, cell in enumerate(row):
            ws[get_column_letter(i + 1) + str(j + 1)].alignment = Alignment(vertical="center", horizontal="center")

            if cell is None:
                continue

            if isinstance(cell, float) or isinstance(cell, int):
                continue

            if isinstance(cell, Decimal):
                continue

            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = (column_width + 2) * 1.75





if __name__ == '__main__':
    wb = createExcel()
    rows = [
        ['考试', '初中预备', '初一期末', '初二期末', '初三期末', ],
        ['排名', 32, 30, 30, 27, ],
        ['区平均分', 80.28, 79.04, 74.32, 105.88, ],
        ['同类平均分', 81.28, 72.04, 73.32, 104.88, ],
    ]
    ws = creatSheet(wb, True, "测试")
    putRowsToSheet(rows, ws)
    wb.save("test.xlsx")
